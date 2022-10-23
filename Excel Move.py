import openpyxl as xl
import os
import shutil

'''
The function copies the template and name the copies with the name of the old files
'''
# parent: the directory contains both the template, new and old folders
# template: the template of new data formate
# old_folder: the directory containing the old files
# new_folder: the directory containing the new files
# indicator: the name to add as an indicator for the new files
def template_copy(parent, template, old_folder, new_folder, indicator):
    # Create the directory of both folder
    old_path = os.path.join(parent, old_folder)
    new_path = os.path.join(parent, new_folder)

    # Iterate over all files in the old folder
    for (root, dirs, files) in os.walk(old_path):
        for names in files:
            # Add the indicator name
            desire_name = names.replace(".xlsx", indicator + ".xlsx")

            # Create the path and copy the files
            new_place = os.path.join(new_path, desire_name)
            file_name = os.path.join(parent, template)
            shutil.copy(file_name, new_place)


'''
'''
# parent: the directory contains both the template, new and old folders
# old_folder: the directory containing the old files
# new_folder: the directory containing the new files
# instruction: the dictionary representing the rows to copy
# indicator: the name to add as an indicator for the new files
# skip: how many lines to skip during copy
def excel_move(parent, old_folder, new_folder, instruction, indicator, skip):
    # Create the directory of both folder
    old_path = os.path.join(parent, old_folder)
    new_path = os.path.join(parent, new_folder)

    for (root, dirs, files) in os.walk(old_path):
        for name in files:
            # Create the path of new and old files
            new_name = name.replace(".xlsx", indicator + ".xlsx")
            old_loc = os.path.join(old_path, name)
            new_loc = os.path.join(new_path, new_name)

            # opening the Excel file
            wb1 = xl.load_workbook(old_loc)
            wb2 = xl.load_workbook(new_loc)

            for content in instruction:
                # Open the desired Excel Sheet
                ws1 = wb1[content[3]]
                ws2 = wb2[content[3]]
                # Find the max column
                mc = ws1.max_column

                for i in range(len(content[0])):
                    for j in range(skip, mc + 1):
                        # Copy and past the style and value as desired
                        to_cell = ws2.cell(j + 1, content[1][i])
                        from_cell = ws1.cell(j + 1, content[0][i])

                        to_cell.value = from_cell.value
                        if content[2]:
                            to_cell.style = from_cell.style
                            to_cell.number_format = from_cell.number_format

            # saving the destination Excel file
            wb2.save(new_loc)


if __name__ == "__main__":
    parent_dir = "C:\\Users\\10648\\OneDrive\\NCPR Intern\\Excel Move"
    original_folder = "Old"
    template = "template.xlsx"
    target_folder = "New"
    indicator = ""
    sheet1 = ([1, 2],
              [1, 2],
              False, "Sheet1")
    sheet1_2 = ([3, 4, 5],
              [3, 4, 5],
              True, "Sheet1")
    sheet2 = ([1, 2, 3, 4, 5],
              [1, 2, 4, 5, 6],
              False, "Sheet2")
    folder_dict = [sheet1, sheet1_2, sheet2, ]
    # Skip the first serval lines
    skip = 2
    template_copy(parent_dir, template, original_folder, target_folder, indicator)
    excel_move(parent_dir, original_folder, target_folder, folder_dict, indicator, skip)
